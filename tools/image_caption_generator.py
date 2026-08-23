#!/usr/bin/env python3
"""从 Excel 读取图片，调用 qwen3.7-plus 生成结构化 Caption。"""

from __future__ import annotations

import argparse
import base64
import binascii
import getpass
import io
import json
import logging
import os
import re
import ssl
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, UnidentifiedImageError


MODEL_ID = "qwen3.7-plus"
BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
MAX_IMAGE_BYTES = 20 * 1024 * 1024
HTTP_TIMEOUT_SECONDS = 30
MAX_MODEL_ATTEMPTS = 3

DEFAULT_CAPTION_RULE = (
    "按照【主体描述】+【修饰词】+【细节补充】+【风格/艺术形式】的逻辑，"
    "整合成一段自然、连贯、客观的中文描述；不要输出栏目标签；控制在80至121字。"
)

SYSTEM_INSTRUCTION = """你是专业的图像 Caption 标注助手。
请只描述图片中能够直接观察到的内容，包括主体、动作、场景、构图、色彩、材质、光线和艺术形式。
不得猜测人物姓名、具体身份、精确地点、事件背景或图片外的信息；不确定的信息应使用审慎措辞。
最终响应必须严格遵循 API 提供的 JSON Schema。"""

SUPPORTED_MIME_BY_FORMAT = {
    "JPEG": "image/jpeg",
    "PNG": "image/png",
    "WEBP": "image/webp",
    "GIF": "image/gif",
    "BMP": "image/bmp",
    "TIFF": "image/tiff",
}


def configure_redirected_console_encoding() -> None:
    """让Windows下被重定向的日志/终端输出保持UTF-8，交互控制台沿用系统设置。"""
    if os.name != "nt":
        return
    for stream in (sys.stdout, sys.stderr):
        if not stream.isatty() and hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")


class UserQuit(Exception):
    """用户要求安全退出。"""


class ImageLoadError(Exception):
    """当前行的所有图片来源均不可用。"""


class ModelCallError(Exception):
    """模型调用或模型结果验证失败。"""

    def __init__(self, message: str, status_code: Optional[int] = None):
        super().__init__(message)
        self.status_code = status_code


@dataclass
class ExcelRow:
    row_number: int
    image_id: str
    source_b: str
    source_c: str
    existing_caption: str


@dataclass
class ImagePayload:
    data_uri: str
    source_type: str
    source_reference: str
    mime_type: str
    width: int
    height: int
    byte_size: int


@dataclass
class WorkbookScan:
    workbook: Workbook
    worksheet: Worksheet
    rows: list[ExcelRow]
    blank_rows: list[int]
    missing_id_rows: list[int]
    duplicate_ids: dict[str, list[int]]


def cell_text(value: Any) -> str:
    """把 Excel 单元格转换为稳定文本，避免整数 ID 变成 1212.0。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def scan_workbook(excel_path: Path) -> WorkbookScan:
    suffix = excel_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("仅支持 .xlsx 和 .xlsm 文件")
    if not excel_path.is_file():
        raise FileNotFoundError(f"文件不存在：{excel_path}")

    workbook = load_workbook(
        excel_path,
        read_only=False,
        data_only=False,
        keep_links=True,
        keep_vba=suffix == ".xlsm",
    )
    if not workbook.worksheets:
        raise ValueError("工作簿中没有工作表")

    worksheet = workbook.worksheets[0]
    rows: list[ExcelRow] = []
    blank_rows: list[int] = []

    for row_number in range(2, worksheet.max_row + 1):
        values = [worksheet.cell(row_number, col).value for col in range(1, 5)]
        if all(value is None or cell_text(value) == "" for value in values):
            blank_rows.append(row_number)
            continue
        rows.append(
            ExcelRow(
                row_number=row_number,
                image_id=cell_text(values[0]),
                source_b=cell_text(values[1]),
                source_c=cell_text(values[2]),
                existing_caption=cell_text(values[3]),
            )
        )

    missing_id_rows = [row.row_number for row in rows if not row.image_id]
    id_rows: dict[str, list[int]] = {}
    for row in rows:
        if row.image_id:
            id_rows.setdefault(row.image_id, []).append(row.row_number)
    duplicate_ids = {
        image_id: row_numbers
        for image_id, row_numbers in id_rows.items()
        if len(row_numbers) > 1
    }

    return WorkbookScan(
        workbook=workbook,
        worksheet=worksheet,
        rows=rows,
        blank_rows=blank_rows,
        missing_id_rows=missing_id_rows,
        duplicate_ids=duplicate_ids,
    )


def print_workbook_summary(scan: WorkbookScan, excel_path: Path) -> None:
    headers = [cell_text(scan.worksheet.cell(1, col).value) for col in range(1, 5)]
    print("\nExcel 读取结果")
    print(f"  文件：{excel_path}")
    print(f"  第一个工作表：{scan.worksheet.title}")
    print(f"  标题：A={headers[0]!r}，B={headers[1]!r}，C={headers[2]!r}，D={headers[3]!r}")
    print(f"  数据行：{len(scan.rows)}")
    print(f"  B列有值：{sum(bool(row.source_b) for row in scan.rows)}")
    print(f"  C列有值：{sum(bool(row.source_c) for row in scan.rows)}")
    print(f"  D列已有内容：{sum(bool(row.existing_caption) for row in scan.rows)}")
    print(f"  完全空白且已忽略的行：{len(scan.blank_rows)}")
    print(f"  缺少图片ID的行：{scan.missing_id_rows or '无'}")
    print(f"  重复图片ID：{scan.duplicate_ids or '无'}")


def prompt_choice(prompt: str, allowed: set[str], default: Optional[str] = None) -> str:
    allowed_upper = {item.upper() for item in allowed}
    while True:
        answer = input(prompt).strip().upper()
        if not answer and default:
            return default.upper()
        if answer in allowed_upper:
            return answer
        print(f"请输入以下选项之一：{'/'.join(sorted(allowed_upper))}")


def choose_excel_path(cli_path: Optional[str]) -> tuple[Path, WorkbookScan]:
    candidate = cli_path
    while True:
        if not candidate:
            candidate = input("请输入 Excel 文件的绝对路径：").strip().strip('"')
        path = Path(candidate).expanduser()
        try:
            path = path.resolve(strict=True)
            scan = scan_workbook(path)
            return path, scan
        except Exception as exc:
            print(f"Excel 读取失败：{exc}")
            if cli_path:
                raise
            if prompt_choice("[R] 重新输入路径  [Q] 退出：", {"R", "Q"}) == "Q":
                raise UserQuit
            candidate = None


def validate_image_bytes(data: bytes) -> tuple[str, int, int]:
    if not data:
        raise ImageLoadError("图片数据为空")
    if len(data) > MAX_IMAGE_BYTES:
        raise ImageLoadError(f"图片超过 {MAX_IMAGE_BYTES // 1024 // 1024} MB 限制")
    try:
        with Image.open(io.BytesIO(data)) as image:
            image_format = (image.format or "").upper()
            width, height = image.size
            image.verify()
    except (UnidentifiedImageError, OSError, ValueError) as exc:
        raise ImageLoadError(f"文件不是可识别的图片：{exc}") from exc

    mime_type = SUPPORTED_MIME_BY_FORMAT.get(image_format)
    if not mime_type:
        raise ImageLoadError(f"暂不支持图片格式：{image_format or '未知'}")
    if width <= 0 or height <= 0:
        raise ImageLoadError("图片尺寸无效")
    return mime_type, width, height


def make_payload(data: bytes, source_type: str, source_reference: str) -> ImagePayload:
    mime_type, width, height = validate_image_bytes(data)
    encoded = base64.b64encode(data).decode("ascii")
    return ImagePayload(
        data_uri=f"data:{mime_type};base64,{encoded}",
        source_type=source_type,
        source_reference=source_reference,
        mime_type=mime_type,
        width=width,
        height=height,
        byte_size=len(data),
    )


def load_data_uri(value: str) -> ImagePayload:
    match = re.fullmatch(
        r"data:(image/[A-Za-z0-9.+-]+);base64,(.+)", value.strip(), flags=re.DOTALL
    )
    if not match:
        raise ImageLoadError("B列不是有效的 data:image/...;base64,... 格式")
    try:
        compact_base64 = re.sub(r"\s+", "", match.group(2))
        data = base64.b64decode(compact_base64, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise ImageLoadError(f"Base64 解码失败：{exc}") from exc
    return make_payload(data, "data_uri", "B列 Data URI")


def read_limited_http_body(response: Any) -> bytes:
    content_length = response.headers.get("Content-Length")
    if content_length:
        try:
            if int(content_length) > MAX_IMAGE_BYTES:
                raise ImageLoadError("URL 图片超过 20 MB 限制")
        except ValueError:
            pass

    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = response.read(64 * 1024)
        if not chunk:
            break
        total += len(chunk)
        if total > MAX_IMAGE_BYTES:
            raise ImageLoadError("URL 图片下载后超过 20 MB 限制")
        chunks.append(chunk)
    return b"".join(chunks)


def load_http_url(url: str) -> ImagePayload:
    if not re.match(r"^https?://", url, flags=re.IGNORECASE):
        raise ImageLoadError("B列只接受 HTTP、HTTPS 或 Data URI")
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 ImageCaptionGenerator/1.0",
            "Accept": "image/avif,image/webp,image/png,image/jpeg,image/gif,image/*;q=0.8,*/*;q=0.1",
        },
    )
    try:
        with urllib.request.urlopen(
            request,
            timeout=HTTP_TIMEOUT_SECONDS,
            context=ssl.create_default_context(),
        ) as response:
            status = getattr(response, "status", 200)
            if status < 200 or status >= 300:
                raise ImageLoadError(f"URL 返回 HTTP {status}")
            data = read_limited_http_body(response)
    except urllib.error.HTTPError as exc:
        raise ImageLoadError(f"URL 返回 HTTP {exc.code}") from exc
    except urllib.error.URLError as exc:
        raise ImageLoadError(f"URL 访问失败：{exc.reason}") from exc
    except (TimeoutError, OSError) as exc:
        raise ImageLoadError(f"URL 访问失败：{exc}") from exc
    return make_payload(data, "http_url", url)


def load_local_path(value: str) -> ImagePayload:
    path = Path(value.strip().strip('"')).expanduser()
    if not path.is_absolute():
        raise ImageLoadError("C列必须是本地绝对路径")
    if not path.is_file():
        raise ImageLoadError(f"本地图片不存在：{path}")
    try:
        file_size = path.stat().st_size
        if file_size > MAX_IMAGE_BYTES:
            raise ImageLoadError("本地图片超过 20 MB 限制")
        data = path.read_bytes()
    except OSError as exc:
        raise ImageLoadError(f"本地图片读取失败：{exc}") from exc
    return make_payload(data, "local_path", str(path))


def resolve_image(source_b: str, source_c: str) -> ImagePayload:
    errors: list[str] = []
    if source_b:
        try:
            if source_b.lower().startswith("data:image/"):
                return load_data_uri(source_b)
            return load_http_url(source_b)
        except ImageLoadError as exc:
            errors.append(f"B列失败：{exc}")
    else:
        errors.append("B列为空")

    if source_c:
        try:
            return load_local_path(source_c)
        except ImageLoadError as exc:
            errors.append(f"C列失败：{exc}")
    else:
        errors.append("C列为空")
    raise ImageLoadError("；".join(errors))


def reload_source_cells(excel_path: Path, row_number: int) -> tuple[str, str]:
    suffix = excel_path.suffix.lower()
    workbook = load_workbook(
        excel_path,
        read_only=True,
        data_only=False,
        keep_links=True,
        keep_vba=suffix == ".xlsm",
    )
    try:
        worksheet = workbook.worksheets[0]
        return (
            cell_text(worksheet.cell(row_number, 2).value),
            cell_text(worksheet.cell(row_number, 3).value),
        )
    finally:
        workbook.close()


def build_json_schema(image_id: str) -> dict[str, Any]:
    return {
        "type": "json_schema",
        "json_schema": {
            "name": "image_caption_result",
            "strict": True,
            "schema": {
                "type": "object",
                "properties": {
                    "image_id": {
                        "type": "string",
                        "enum": [image_id],
                        "description": "当前图片的唯一ID，必须原样返回",
                    },
                    "image_info": {
                        "type": "string",
                        "description": "对主体、动作、场景等可见内容的客观语义概括",
                    },
                    "caption": {
                        "type": "string",
                        "description": "按照用户描述规范生成的最终Caption",
                    },
                },
                "required": ["image_id", "image_info", "caption"],
                "additionalProperties": False,
            },
        },
    }


def create_client(api_key: str) -> Any:
    try:
        from openai import OpenAI
    except ImportError as exc:
        raise RuntimeError(
            "缺少 openai 包，请先运行：pip install -r requirements_caption.txt"
        ) from exc
    return OpenAI(
        api_key=api_key,
        base_url=BASE_URL,
        timeout=90.0,
        max_retries=0,
    )


def caption_character_count(caption: str) -> int:
    """忽略空白字符计数，中文标点计入。"""
    return len(re.sub(r"\s+", "", caption))


def object_for_log(value: Any) -> Any:
    """将 SDK 对象转换为可完整写入 JSON 日志的普通对象。"""
    if hasattr(value, "model_dump"):
        try:
            return value.model_dump(mode="json")
        except TypeError:
            return value.model_dump()
    if isinstance(value, dict):
        return {key: object_for_log(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [object_for_log(item) for item in value]
    return value


def log_model_payload(
    label: str,
    image_id: str,
    attempt: int,
    payload: Any,
    level: int = logging.INFO,
) -> None:
    """完整记录单次模型请求、响应或错误；API Key 不属于这些载荷。"""
    rendered = json.dumps(
        object_for_log(payload),
        ensure_ascii=False,
        indent=2,
        default=str,
    )
    logging.log(
        level,
        "%s image_id=%s attempt=%d\n%s",
        label,
        image_id,
        attempt,
        rendered,
    )


def parse_completion(completion: Any, expected_image_id: str) -> dict[str, str]:
    if not getattr(completion, "choices", None):
        raise ModelCallError("API 响应中没有 choices")
    message = completion.choices[0].message
    refusal = getattr(message, "refusal", None)
    if refusal:
        raise ModelCallError(f"模型拒绝处理：{refusal}")
    content = getattr(message, "content", None)
    if not content:
        raise ModelCallError("模型没有返回结构化内容")
    try:
        result = json.loads(content)
    except json.JSONDecodeError as exc:
        raise ModelCallError(f"模型结果不是有效 JSON：{exc}") from exc

    if isinstance(result, list):
        if len(result) == 1 and isinstance(result[0], dict):
            logging.warning(
                "MODEL_SCHEMA_SERVICE_VIOLATION image_id=%s："
                "API声明根节点为object，但服务端返回了单元素数组；"
                "脚本将移除数组外壳后继续执行完整Schema校验",
                expected_image_id,
            )
            result = result[0]
        else:
            raise ModelCallError(
                "模型结果应为对象，服务端却返回数组："
                f"数组长度={len(result)}；仅单元素对象数组可安全归一化"
            )

    required_keys = {"image_id", "image_info", "caption"}
    if not isinstance(result, dict):
        raise ModelCallError(
            f"模型结果应为对象，实际类型为 {type(result).__name__}"
        )
    actual_keys = set(result)
    if actual_keys != required_keys:
        missing_keys = sorted(required_keys - actual_keys)
        extra_keys = sorted(actual_keys - required_keys)
        raise ModelCallError(
            "模型结果字段与 JSON Schema 不一致："
            f"缺失字段={missing_keys or '无'}，额外字段={extra_keys or '无'}，"
            f"实际字段={sorted(actual_keys)}"
        )
    invalid_types = {
        key: type(result[key]).__name__
        for key in sorted(required_keys)
        if not isinstance(result[key], str)
    }
    if invalid_types:
        raise ModelCallError(
            f"模型结果字段类型与 JSON Schema 不一致：{invalid_types}"
        )
    if result["image_id"] != expected_image_id:
        raise ModelCallError(
            "模型返回的 image_id 与当前行不一致："
            f"期望={expected_image_id!r}，实际={result['image_id']!r}"
        )
    if not result["image_info"].strip() or not result["caption"].strip():
        raise ModelCallError("image_info 或 caption 为空")
    return {key: result[key].strip() for key in required_keys}


def completion_usage(completion: Any) -> dict[str, Any]:
    usage = getattr(completion, "usage", None)
    if usage is None:
        return {}
    if hasattr(usage, "model_dump"):
        return usage.model_dump()
    if isinstance(usage, dict):
        return usage
    return {}


def merge_usage_totals(target: dict[str, Any], source: dict[str, Any]) -> None:
    """累计自动重试产生的Token用量，避免日志只记录最后一次请求。"""
    for key, value in source.items():
        if isinstance(value, bool):
            target[key] = value
        elif isinstance(value, (int, float)):
            target[key] = target.get(key, 0) + value
        elif isinstance(value, dict):
            nested = target.setdefault(key, {})
            if isinstance(nested, dict):
                merge_usage_totals(nested, value)
        elif key not in target:
            target[key] = value


def generate_caption(
    client: Any,
    image_id: str,
    image: ImagePayload,
    caption_rule: str,
    enforce_default_length: bool,
) -> tuple[dict[str, str], dict[str, Any], str, float]:
    correction = ""
    last_error: Optional[Exception] = None
    total_usage: dict[str, Any] = {}
    overall_started = time.perf_counter()

    for attempt in range(1, MAX_MODEL_ATTEMPTS + 1):
        user_text = (
            f"当前图片ID：{image_id}\n"
            f"Caption描述规范：{caption_rule}\n"
            "请先在 image_info 中客观归纳画面，再在 caption 中给出最终可供评测人员直接使用或改写的描述。"
        )
        if correction:
            user_text += f"\n上一次结果需要修正：{correction}"

        request_payload = {
            "model": MODEL_ID,
            "messages": [
                {"role": "system", "content": SYSTEM_INSTRUCTION},
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {"url": image.data_uri},
                        },
                        {"type": "text", "text": user_text},
                    ],
                },
            ],
            "response_format": build_json_schema(image_id),
            "extra_body": {"enable_thinking": False},
        }
        log_model_payload("MODEL_INPUT", image_id, attempt, request_payload)

        try:
            completion = client.chat.completions.create(**request_payload)
            log_model_payload("MODEL_OUTPUT", image_id, attempt, completion)
            merge_usage_totals(total_usage, completion_usage(completion))
            result = parse_completion(completion, image_id)

            if enforce_default_length:
                length = caption_character_count(result["caption"])
                if not 80 <= length <= 121:
                    correction = (
                        f"caption 当前为 {length} 字。请改写到80至121字，"
                        "保持客观、自然，不要改变JSON字段。"
                    )
                    last_error = ModelCallError(correction)
                    if attempt < MAX_MODEL_ATTEMPTS:
                        logging.warning(
                            "图片ID=%s Caption长度不合格，准备第%d次重试",
                            image_id,
                            attempt + 1,
                        )
                        continue
                    raise last_error

            return (
                result,
                total_usage,
                str(getattr(completion, "id", "")),
                time.perf_counter() - overall_started,
            )
        except ModelCallError as exc:
            last_error = exc
            logging.warning(
                "MODEL_VALIDATION_ERROR image_id=%s attempt=%d：%s",
                image_id,
                attempt,
                exc,
            )
            if attempt < MAX_MODEL_ATTEMPTS:
                correction = str(exc)
                time.sleep(min(2**attempt, 8))
                continue
            break
        except Exception as exc:
            error_payload = {
                "exception_type": exc.__class__.__name__,
                "message": str(exc),
                "status_code": getattr(exc, "status_code", None),
                "request_id": getattr(exc, "request_id", None),
                "body": getattr(exc, "body", None),
            }
            log_model_payload(
                "MODEL_OUTPUT_ERROR",
                image_id,
                attempt,
                error_payload,
                level=logging.ERROR,
            )
            last_error = exc
            status_code = getattr(exc, "status_code", None)
            transient = (
                status_code in {408, 409, 429}
                or isinstance(status_code, int)
                and status_code >= 500
                or exc.__class__.__name__ in {"APIConnectionError", "APITimeoutError"}
            )
            if transient and attempt < MAX_MODEL_ATTEMPTS:
                logging.warning(
                    "图片ID=%s API临时错误，%d秒后重试：%s",
                    image_id,
                    min(2**attempt, 8),
                    exc,
                )
                time.sleep(min(2**attempt, 8))
                continue
            raise ModelCallError(str(exc), status_code=status_code) from exc

    if isinstance(last_error, ModelCallError):
        raise last_error
    raise ModelCallError(str(last_error or "模型调用失败"))


def atomic_save(workbook: Workbook, output_path: Path) -> None:
    temporary = output_path.with_name(
        f".{output_path.stem}.saving{output_path.suffix}"
    )
    try:
        workbook.save(temporary)
        os.replace(temporary, output_path)
    finally:
        if temporary.exists():
            temporary.unlink(missing_ok=True)


def jsonl_write(handle: Any, payload: dict[str, Any]) -> None:
    handle.write(json.dumps(payload, ensure_ascii=False) + "\n")
    handle.flush()


def safe_source_reference(image: ImagePayload) -> str:
    if image.source_type == "data_uri":
        return "B列 Data URI（内容未写入日志）"
    if image.source_type == "http_url":
        parts = urllib.parse.urlsplit(image.source_reference)
        return urllib.parse.urlunsplit(
            (parts.scheme, parts.netloc, parts.path, "", "")
        )
    return image.source_reference


def make_output_paths(excel_path: Path) -> tuple[Path, Path, Path]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = excel_path.parent / "caption_output"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_excel = output_dir / f"{excel_path.stem}_caption结果_{timestamp}{excel_path.suffix}"
    log_path = output_dir / f"{excel_path.stem}_caption_{timestamp}.log"
    jsonl_path = output_dir / f"{excel_path.stem}_caption_{timestamp}.jsonl"
    return output_excel, log_path, jsonl_path


def configure_logging(log_path: Path) -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
        force=True,
    )


def run_check_only(scan: WorkbookScan) -> int:
    succeeded = 0
    failed = 0
    print("\n开始检查图片来源（不会调用模型，也不会修改Excel）")
    for row in scan.rows:
        if not row.image_id:
            print(f"  第{row.row_number}行：失败，图片ID为空")
            failed += 1
            continue
        try:
            image = resolve_image(row.source_b, row.source_c)
            print(
                f"  第{row.row_number}行 ID={row.image_id}：成功，"
                f"来源={image.source_type}，{image.mime_type}，"
                f"{image.width}x{image.height}，{image.byte_size}字节"
            )
            succeeded += 1
        except ImageLoadError as exc:
            print(f"  第{row.row_number}行 ID={row.image_id or '<空>'}：失败，{exc}")
            failed += 1
    print(f"\n检查完成：成功 {succeeded}，失败 {failed}")
    return 0 if failed == 0 else 2


def process_workbook(excel_path: Path, scan: WorkbookScan) -> int:
    if scan.missing_id_rows or scan.duplicate_ids:
        print("\n存在空ID或重复ID。请先修正Excel，脚本不会调用模型。")
        return 2

    if prompt_choice("\n以上内容是否读取正确？[Y/N]（默认Y）：", {"Y", "N"}, "Y") == "N":
        print("已取消，未修改文件。")
        return 0

    print("\n默认 Caption 规范：")
    print(DEFAULT_CAPTION_RULE)
    custom_rule = input("输入自定义规范，或直接按回车使用默认规范：").strip()
    caption_rule = custom_rule or DEFAULT_CAPTION_RULE
    enforce_default_length = not bool(custom_rule)

    api_key = ""
    while not api_key:
        api_key = getpass.getpass("请输入阿里云百炼 API Key（输入内容不会显示）：").strip()
        if not api_key:
            print("API Key 不能为空。")

    output_excel, log_path, jsonl_path = make_output_paths(excel_path)
    configure_logging(log_path)
    logging.info("开始处理文件：%s", excel_path)
    logging.info("模型：%s", MODEL_ID)
    logging.info("Base URL：%s", BASE_URL)
    logging.info("输出文件：%s", output_excel)

    client = create_client(api_key)
    atomic_save(scan.workbook, output_excel)
    counters = {"success": 0, "skipped": 0, "failed": 0}

    with jsonl_path.open("a", encoding="utf-8") as jsonl_handle:
        try:
            for index, row in enumerate(scan.rows, 1):
                print(f"\n[{index}/{len(scan.rows)}] 第{row.row_number}行，图片ID={row.image_id}")

                existing = cell_text(scan.worksheet.cell(row.row_number, 4).value)
                if existing:
                    choice = prompt_choice(
                        "D列已有内容：[O] 覆盖  [S] 跳过  [Q] 保存并退出：",
                        {"O", "S", "Q"},
                    )
                    if choice == "S":
                        counters["skipped"] += 1
                        logging.info("跳过第%d行 ID=%s：D列已有内容", row.row_number, row.image_id)
                        jsonl_write(
                            jsonl_handle,
                            {
                                "timestamp": datetime.now().isoformat(timespec="seconds"),
                                "status": "skipped_existing_caption",
                                "row_number": row.row_number,
                                "image_id": row.image_id,
                            },
                        )
                        continue
                    if choice == "Q":
                        raise UserQuit

                source_b, source_c = row.source_b, row.source_c
                while True:
                    try:
                        image = resolve_image(source_b, source_c)
                        break
                    except ImageLoadError as exc:
                        logging.error("第%d行 ID=%s 图片获取失败：%s", row.row_number, row.image_id, exc)
                        choice = prompt_choice(
                            f"图片获取失败：{exc}\n[R] 修改并保存源Excel后重新检查  "
                            "[S] 跳过  [Q] 保存并退出：",
                            {"R", "S", "Q"},
                        )
                        if choice == "R":
                            try:
                                source_b, source_c = reload_source_cells(excel_path, row.row_number)
                                scan.worksheet.cell(row.row_number, 2).value = source_b or None
                                scan.worksheet.cell(row.row_number, 3).value = source_c or None
                                print("已重新读取当前行的B、C列。")
                            except Exception as reload_exc:
                                print(f"重新读取失败：{reload_exc}")
                            continue
                        if choice == "S":
                            counters["skipped"] += 1
                            jsonl_write(
                                jsonl_handle,
                                {
                                    "timestamp": datetime.now().isoformat(timespec="seconds"),
                                    "status": "skipped_image_error",
                                    "row_number": row.row_number,
                                    "image_id": row.image_id,
                                    "error": str(exc),
                                },
                            )
                            image = None
                            break
                        raise UserQuit

                if image is None:
                    continue

                while True:
                    try:
                        result, usage, response_id, elapsed = generate_caption(
                            client,
                            row.image_id,
                            image,
                            caption_rule,
                            enforce_default_length,
                        )
                        break
                    except ModelCallError as exc:
                        logging.error("第%d行 ID=%s 模型调用失败：%s", row.row_number, row.image_id, exc)
                        choice = prompt_choice(
                            f"模型调用失败：{exc}\n[R] 重试  [K] 重新输入Key  "
                            "[S] 跳过  [Q] 保存并退出：",
                            {"R", "K", "S", "Q"},
                        )
                        if choice == "R":
                            continue
                        if choice == "K":
                            api_key = getpass.getpass("请输入新的 API Key：").strip()
                            if not api_key:
                                print("API Key 不能为空。")
                                continue
                            client = create_client(api_key)
                            continue
                        if choice == "S":
                            counters["failed"] += 1
                            jsonl_write(
                                jsonl_handle,
                                {
                                    "timestamp": datetime.now().isoformat(timespec="seconds"),
                                    "status": "model_error",
                                    "row_number": row.row_number,
                                    "image_id": row.image_id,
                                    "source_type": image.source_type,
                                    "source_reference": safe_source_reference(image),
                                    "error": str(exc),
                                },
                            )
                            result = None
                            break
                        raise UserQuit

                if result is None:
                    continue

                scan.worksheet.cell(row.row_number, 4).value = result["caption"]
                atomic_save(scan.workbook, output_excel)
                counters["success"] += 1
                logging.info(
                    "成功 第%d行 ID=%s 来源=%s 尺寸=%dx%d 耗时=%.2fs",
                    row.row_number,
                    row.image_id,
                    image.source_type,
                    image.width,
                    image.height,
                    elapsed,
                )
                jsonl_write(
                    jsonl_handle,
                    {
                        "timestamp": datetime.now().isoformat(timespec="seconds"),
                        "status": "success",
                        "row_number": row.row_number,
                        "image_id": row.image_id,
                        "source_type": image.source_type,
                        "source_reference": safe_source_reference(image),
                        "image": {
                            "mime_type": image.mime_type,
                            "width": image.width,
                            "height": image.height,
                            "byte_size": image.byte_size,
                        },
                        "response_id": response_id,
                        "elapsed_seconds": round(elapsed, 3),
                        "usage": usage,
                        "result": result,
                    },
                )
        except UserQuit:
            logging.info("用户要求退出，正在保存当前进度。")
        finally:
            atomic_save(scan.workbook, output_excel)
            scan.workbook.close()

    print("\n处理结束")
    print(f"  成功：{counters['success']}")
    print(f"  跳过：{counters['skipped']}")
    print(f"  失败：{counters['failed']}")
    print(f"  Excel结果：{output_excel}")
    print(f"  运行日志：{log_path}")
    print(f"  结构化日志：{jsonl_path}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="读取Excel图片并调用 qwen3.7-plus 生成 Caption"
    )
    parser.add_argument(
        "excel_path",
        nargs="?",
        help="可选。Excel绝对路径；省略时由脚本交互询问。",
    )
    parser.add_argument(
        "--check-only",
        action="store_true",
        help="只检查Excel和图片来源，不询问API Key、不调用模型、不修改文件。",
    )
    return parser


def main() -> int:
    configure_redirected_console_encoding()
    args = build_parser().parse_args()
    try:
        excel_path, scan = choose_excel_path(args.excel_path)
        print_workbook_summary(scan, excel_path)
        try:
            if args.check_only:
                return run_check_only(scan)
            return process_workbook(excel_path, scan)
        finally:
            scan.workbook.close()
    except UserQuit:
        print("已退出，未开始新的处理。")
        return 0
    except KeyboardInterrupt:
        print("\n用户中断。")
        return 130
    except Exception as exc:
        print(f"运行失败：{exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
